from openpyxl import load_workbook
import pandas as pd
import numpy as np
import datetime
import random
import string
import json
import pyodbc
import openpyxl


class DB_Updates():

    def Retrieve_dropdown(self):
        conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_str)

        table_name = "MF_Entity_Mapping"
        # SQL query to select data from the table
        select_query = f"SELECT * FROM {table_name}"

        # Retrieve data using pandas
        df = pd.read_sql_query(select_query, conn)

        # Close the connection
        conn.close()
        return(df)

    def Generate_DatasetUID(self,input_data_dictionary):
        #extracted_json = json.loads(json_ui)

        prefix = 'SYN'
        random_prefix = "".join(random.sample(string.digits, 6))
        DatasetUID = input_data_dictionary['DataSetName']+prefix + random_prefix
        return(DatasetUID)

    def Get_DB_Status_values(self):
        conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_str)

        # Create a cursor to execute SQL queries
        cursor = conn.cursor()

        # Update a column value in a row
        table_name = "DB_Status"
        select_query = f"SELECT * FROM {table_name}"

        # Execute the SELECT query
        cursor.execute(select_query)
        rows = cursor.fetchall()

        status_dictionary = dict()

        for row in rows:
            status_dictionary[int(row[0])] = row[1]

        # conn.commit()
        print(status_dictionary)
        print("Record updated successfully!")

        # Close the connection and cursor
        cursor.close()
        conn.close()
        return(status_dictionary)

    def Insert_Dataset(self,json_ui):
        status = 'Draft'
        now = datetime.datetime.now()
        now =now.isoformat()
        conn_string = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_string)

        cursor = conn.cursor()
        table_name = "DataSet"
        VALUES = ({json_ui['ProcessAreaId']}, {json_ui['TargetSys']}, {json_ui['DatasetUID']}, {json_ui['DataSetName']},
                  {now}, {json_ui['CreatedBy']}, {status})

        VALUES = (json_ui['ProcessAreaId'], json_ui['TargetSys'], json_ui['DatasetUID'], json_ui['DataSetName'],
                  now, json_ui['CreatedBy'], status)
        insert_query = f"""INSERT INTO {table_name} ([Process_Area], [Target_system], [DataSet_GUID], [DataSet_Name], [Created_On], [Created_By], [Status]) VALUES (?, ?, ?, ?, ?, ?, ?)"""
        cursor.execute(insert_query, VALUES)
        conn.commit()
        print("Records inserted successfully!")
        # Close the connection and cursor/
        cursor.close()
        conn.close()

    def Update_DataSet(self,json_ui,StatusID):
        print('Update_DataSet')
        # Connect to the Access database
        conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_str)
        # Create a cursor to execute SQL queries
        cursor = conn.cursor()
        # Update a column value in a row
        table_name = "DataSet"
        column_to_update = "Changed_On"
        status_value = StatusID#"Draft Generation in Progress "
        new_value2 = 'vamsi'
        condition_column = "DataSet_GUID"
        condition_value = json_ui['DatasetUID']
        now = datetime.datetime.now()
        now = now.isoformat()
        update_query = f"UPDATE {table_name} SET Changed_On = ?, Status = ? WHERE {condition_column} = ?"
        cursor.execute(update_query, (now, status_value, condition_value))
        conn.commit()
        print("Record updated successfully!")

        # Close the connection and cursor
        cursor.close()
        conn.close()


    def Update_DataSet_Log(self,json_ui,concat_dataframe):
        print("Update_dataset_log")
        concat_dataframe['DataSet_GUID'] = json_ui['DatasetUID']
        concat_dataframe['DataSet_Name'] = json_ui['DataSetName']

        filename = r'C:\Users\vamsikkrishna\PycharmProjects\pythonProject1\DB_FILES\DataSet_Log.xlsx'
        sheet_name = 'Sheet1'
        #k = self.append_df_to_excel(filename,concat_dataframe,sheet_name='Sheet1')
        workbook = openpyxl.load_workbook(filename)

        sheet = workbook[sheet_name]

        # Get the last row index in the sheet
        last_row = sheet.max_row
        print('last row data', last_row)

        for col_idx, col_name in enumerate(concat_dataframe.columns, start=1):
            sheet.cell(row=1, column=col_idx, value=col_name)

        for idx, row in concat_dataframe.iterrows():
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=last_row + idx, column=col_idx, value=value)

        # Save the workbook
        workbook.save(filename)

        print("datset_log_updated")



        #concat_dataframe.to_excel(r'C:\Users\vamsikkrishna\PycharmProjects\pythonProject1\DB_FILES\DataSet_Log.xlsx',index=False)
        print('Dataset_log_updated')

    def Database_Input_update(self,json_ui):
        print('Database Input Update')

        now = datetime.datetime.now()
        now = now.isoformat()
        conn_string = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_string)

        cursor = conn.cursor()
        table_name = "DataSet"
        VALUES = ({json_ui['ProcessAreaId']}, {json_ui['TargetSys']}, {json_ui['DatasetUID']}, {json_ui['DataSetName']},
                  {now}, {json_ui['CreatedBy']})

        VALUES = (json_ui['ProcessAreaId'], json_ui['TargetSys'], json_ui['DatasetUID'], json_ui['DataSetName'],
                  now, json_ui['CreatedBy'])
        insert_query = f"""INSERT INTO {table_name} ([Process_Area], [Target_system], [DataSet_GUID], [DataSet_Name], [Created_On], [Created_By], [Status]) VALUES (?, ?, ?, ?, ?, ?, ?)"""
        cursor.execute(insert_query, VALUES)
        conn.commit()
        print("Records inserted successfully!")
        # Close the connection and cursor
        cursor.close()
        conn.close()

    def Input_Draft_Data(self,header,body_list):
        print('Input_Draft_Data')
        print(body_list)
        temp = pd.DataFrame.from_records(body_list)

        current_datetime = datetime.datetime.now()

        # Extract only the date and format it as 'mm-dd-yyyy'
        formatted_date = current_datetime.strftime('%m-%d-%Y')

        # Print the formatted date
        print(formatted_date)

        temp['DatasetUID'] = header['DatasetUID']
        temp['TargetSys'] = header['TargetSys']
        temp['DataSetName'] = header['DataSetName']
        temp['ChangedOn'] = formatted_date
        print(pd.DataFrame.from_records(body_list))
        print(temp.to_json(orient='records'))

        temp.loc[temp['FieldValue'].apply(
            lambda x: isinstance(x, list) and all(isinstance(item, int) for item in x)), 'FieldValue'] = temp.loc[
            temp['FieldValue'].apply(
                lambda x: isinstance(x, list) and all(isinstance(item, int) for item in x)), 'FieldValue'].apply(
            lambda x: ','.join(map(str, x)))
        conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_str)
        crsr = conn.cursor()

        table_name = "DataSet_Input"

        col_values = ", ".join(temp.columns.tolist())
        col_placeholder = ', '.join(['? '] * len(temp.columns.tolist()))

        insert_query = f"INSERT INTO [{table_name}] ({col_values}) VALUES ({col_placeholder})"
        crsr.executemany(
            f"INSERT INTO [{table_name}] ({col_values}) VALUES ({col_placeholder})",
            temp.itertuples(index=False))
        conn.commit()
        crsr.close()
        conn.close()

    def Draft_Values_response(self,header,DB_Status_dictionary):
        conn_str = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\vamsikkrishna\Documents\Database1.accdb;'
        conn = pyodbc.connect(conn_str)
        #crsr = conn.cursor()

        table_name = "DataSet"
        DatasetUID = header['DatasetUID']
        condition_column = 'DataSet_GUID'
        status_id = ''
        select_query = f"SELECT Process_Area,DataSet_Name,DataSet_GUID, Status from {table_name} WHERE {condition_column} = '{DatasetUID}'"

        df = pd.read_sql_query(select_query, conn)
        conn.close()
        print(df)
        status_value = df['Status'].squeeze().strip()
        print(status_value)
        for key, value in DB_Status_dictionary.items():
            # print(key,value)
            print(f"{value}and {status_value}")
            if value.strip() == status_value.strip():
                status_id = key
        # key = list(filter(lambda x: DB_Status_dictionary[x] == df['Status'].values[0], DB_Status_dictionary))[0]
        df['statusID'] = status_id
        print(df)

        df.rename({"Process_Area": 'processAreaId', 'DataSet_Name': 'dataSetName', 'DataSet_GUID': 'dataSetUID',
                   'Status': 'statusDesc'}, inplace=True)

        response = df.to_json(orient='records')
        response_new = response[1:-1]


        return response_new
















